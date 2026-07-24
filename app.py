"""
Ứng dụng Streamlit: Xử lý Bảng Điểm Tổng Hợp
Data flow:
  1. Upload nhiều file Excel bảng điểm (sheet "Điểm tổng hợp") → trích xuất điểm
  2. Chọn chuyên ngành → load KetQua_<chuyên ngành>.xlsx từ thư mục data/
     (Tên môn học, MSMH, Trọng số TX/GK/CK, Điểm đạt)
  3. Join theo Tên môn học (case-insensitive) → tạo Mã lớp học phần
  4. Xuất 1 file Excel 3 sheet: Daydu / LopHocPhan / SV_LopHocPhan
"""
import os, re, io, unicodedata
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Xử lý Bảng Điểm", page_icon="🎓", layout="centered")

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

CHUYEN_NGANH = {
    "BSNT - Ngoại":         "KetQua_BSNT_Ngoai.xlsx",
    "BSNT - Sản":           "KetQua_BSNT_San.xlsx",
    "BSNT - Nhi":           "KetQua_BSNT_Nhi.xlsx",
    "BSNT - Tai Mũi Họng":  "KetQua_BSNT_TMH.xlsx",
    "CKI - Ngoại khoa":     "KetQua_CKI_Ngoai.xlsx",
    "CKI - Sản phụ khoa":   "KetQua_CKI_San.xlsx",
    "CKI - Nhi khoa":       "KetQua_CKI_Nhi.xlsx",
    "CKI - Tai Mũi Họng":   "KetQua_CKI_TMH.xlsx",
    "CKI - Nội khoa":       "KetQua_CKI_Noi.xlsx",
    "CKI - Ngoại Tim Mạch": "KetQua_CKI_NgoaiTimMach.xlsx",
}
HK_MAP = {"I":"1","II":"2","III":"3","1":"1","2":"2","3":"3"}

# ── CSS ─────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Be Vietnam Pro',sans-serif}
.hero{background:linear-gradient(135deg,#1e3a5f,#2563a8);color:#fff;
      padding:1.8rem 2.2rem;border-radius:14px;margin-bottom:1.8rem}
.hero h1{margin:0;font-size:1.65rem;font-weight:700}
.hero p{margin:.35rem 0 0;opacity:.82;font-size:.9rem}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;
      padding:1.3rem 1.5rem;margin-bottom:1.1rem;box-shadow:0 1px 4px rgba(0,0,0,.05)}
.step{font-size:.68rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;
      color:#2563a8;margin-bottom:.35rem}
.ok{background:#f0fdf4;border:1px solid #86efac;border-radius:10px;padding:.9rem 1.3rem;margin:.7rem 0}
.ok h4{margin:0 0 .25rem;color:#166534;font-size:.9rem}
.warn{background:#fffbeb;border:1px solid #fcd34d;border-radius:10px;
      padding:.8rem 1.2rem;margin:.5rem 0;font-size:.84rem;color:#92400e}
.err{background:#fef2f2;border:1px solid #fca5a5;border-radius:10px;
     padding:.8rem 1.2rem;margin:.5rem 0;font-size:.84rem;color:#991b1b}
.miss{background:#f8fafc;border:1px dashed #cbd5e1;border-radius:10px;
      padding:.9rem 1.3rem;font-size:.83rem;color:#475569}
.miss code{background:#e2e8f0;padding:1px 5px;border-radius:4px}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# CORE FUNCTIONS
# ══════════════════════════════════════════════════════════════
def normalize_score(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    try: return round(float(str(v).strip().replace(",",".")),1)
    except: return None

def round_0_1(v):
    """Làm tròn điểm thành phần đến 0.1 (VD: 7.24 -> 7.2, 7.26 -> 7.3)."""
    if v is None: return None
    return round(round(v*10)/10, 1)

def round_0_5(v):
    """Làm tròn điểm trung bình đến 0.5 (VD: 7.24 -> 7.0, 7.26 -> 7.5)."""
    if v is None: return None
    return round(round(v*2)/2, 1)

def normalize_nam_hoc(v):
    """Chuẩn hóa năm học: '2024 - 2025' / '2024– 2025' → '2024-2025'."""
    if not v: return v
    return re.sub(r'\s*[-\u2013\u2014]\s*', '-', str(v).strip())

def normalize_hoc_ky(v):
    """Chuẩn hóa học kỳ: 'I' -> '1', 'II' -> '2', 'III' -> '3'; giữ nguyên nếu đã là số hoặc không nhận diện được."""
    if not v: return v
    key = str(v).strip().upper()
    return HK_MAP.get(key, str(v).strip())

def extract_general_info(ws):
    info={"Khoa":None,"Lớp":None,"Tên môn học":None,"Năm học":None,"Học kỳ":None,"Số tín chỉ":None}
    kws={"khoa":"Khoa","lớp":"Lớp","tên môn học":"Tên môn học",
         "năm học":"Năm học","học kỳ":"Học kỳ","số tín chỉ":"Số tín chỉ"}
    for row in ws.iter_rows(min_row=1,max_row=12):
        rv=[(c.column,str(c.value).strip() if c.value else "") for c in row]
        for i,(_,val) in enumerate(rv):
            for kw,field in kws.items():
                if kw in val.lower() and info[field] is None:
                    for _,nv in rv[i+1:]:
                        if nv and nv not in(":",""): info[field]=nv; break
    info["N\u0103m h\u1ecdc"] = normalize_nam_hoc(info["N\u0103m h\u1ecdc"])
    info["H\u1ecdc k\u1ef3"] = normalize_hoc_ky(info["H\u1ecdc k\u1ef3"])
    return info

def find_header_row(ws):
    for row in ws.iter_rows(min_row=1,max_row=20):
        for cell in row:
            if cell.value and "mssv" in str(cell.value).lower(): return cell.row
    return None

def find_data_start_row(ws,from_row):
    for row in ws.iter_rows(min_row=from_row+1,max_row=from_row+10):
        for cell in row:
            if cell.value is not None:
                try: int(float(str(cell.value).replace(",","."))); return cell.row
                except: pass
    return from_row+2

def map_columns(ws,header_rows):
    col_map={}; all_headers={}
    for hr in header_rows:
        for cell in ws[hr]:
            if cell.value:
                txt=str(cell.value).lower().strip().replace("\n"," ")
                if cell.column not in all_headers: all_headers[cell.column]=txt
    for col,txt in all_headers.items():
        if   "mssv" in txt:                                          col_map["mssv"]=col
        elif txt in("họ","họ và tên","ho"):                         col_map["ho"]=col
        elif txt in("tên","ten"):                                    col_map["ten"]=col
        elif "thường xuyên" in txt or "quá trình" in txt:          col_map["diem_qt"]=col
        elif "giữa kỳ" in txt:           col_map["diem_gk"]=col
        elif "kết thúc" in txt or("thi" in txt and "điểm" in txt):  col_map["diem_thi"]=col
        elif "làm tròn" in txt or "đ. số" in txt:                   col_map["diem_lam_tron"]=col
        elif "điểm tb" in txt or "trung bình" in txt:               col_map["diem_tb"]=col
        elif "đ. chữ" in txt or "bằng chữ" in txt:                 col_map["diem_chu"]=col
    for row in ws.iter_rows(min_row=max(1,min(header_rows)-1),
                            max_row=min(max(header_rows)+3,ws.max_row)):
        for cell in row:
            if cell.value:
                t=str(cell.value).strip()
                if t=="30%" and "diem_qt"  not in col_map: col_map["diem_qt"] =cell.column
                if t=="30%" and "diem_gk"  not in col_map: col_map["diem_gk"] =cell.column
                if t=="70%" and "diem_thi" not in col_map: col_map["diem_thi"]=cell.column
    return col_map

def extract_score_table(ws):
    hr=find_header_row(ws)
    if hr is None: return []
    hrs=[hr]
    for d in [1,2]:
        chk=hr+d
        if chk<=ws.max_row:
            for cell in ws[chk]:
                if cell.value and str(cell.value).strip() in("30%","70%","Đ. Số","Đ. Chữ"):
                    if chk not in hrs: hrs.append(chk)
    cm=map_columns(ws,hrs); ds=find_data_start_row(ws,max(hrs))
    records=[]
    for row in ws.iter_rows(min_row=ds,max_row=ws.max_row):
        rt=" ".join(str(c.value) for c in row if c.value).lower()
        if any(k in rt for k in("số hv","tổng số","sĩ số")): break
        mv=row[cm["mssv"]-1].value if "mssv" in cm else None
        if not mv: continue
        mv=str(mv).strip()
        if mv.lower() in("mssv","nan","none",""): continue
        def get(k): return row[cm[k]-1].value if k in cm else None
        ho =str(get("ho")).strip()  if get("ho")  else ""
        ten=str(get("ten")).strip() if get("ten") else ""
        records.append({"MSSV":mv,"Họ và tên":f"{ho} {ten}".strip(),
            "Điểm quá trình (30%)":round_0_1(normalize_score(get("diem_qt"))),
            "Điểm giữa kỳ (30%)": round_0_1(normalize_score(get("diem_gk"))),
            "Điểm thi (70%)":      round_0_1(normalize_score(get("diem_thi"))),
            "Điểm TB":             normalize_score(get("diem_tb")),
            "Điểm làm tròn":       round_0_5(normalize_score(get("diem_tb"))),
            "Điểm chữ":            str(get("diem_chu")).strip() if get("diem_chu") else ""})
    return records

# ── Xử lý file dạng công thức (DSHVDT + diem_tx + diem_kt) ──

def is_formula(v):
    return isinstance(v, str) and v.startswith("=")

def real_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return None if is_formula(v) else v

def _find_stt_data_start(ws):
    """Tìm (header_row, data_start_row) dựa vào cột STT."""
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and not is_formula(cell.value) and str(cell.value).strip().upper() == "STT":
                hr = cell.row
                for dr in range(hr+1, hr+6):
                    v = real_val(ws, dr, cell.column)
                    if v is not None:
                        try:
                            if int(float(str(v))) >= 1: return hr, dr
                        except: pass
    return None, None

def _extract_info_dshvdt(ws):
    info = {"Khoa":None,"Lớp":None,"Tên môn học":None,
            "Năm học":None,"Học kỳ":None,"Số tín chỉ":None}
    kws  = {"khoa":"Khoa","lớp":"Lớp","tên môn học":"Tên môn học",
            "năm học":"Năm học","học kỳ":"Học kỳ","tín chỉ":"Số tín chỉ"}
    for row in ws.iter_rows(min_row=1, max_row=12):
        rv = [(c.column, str(c.value).strip())
              for c in row if c.value and not is_formula(c.value) and str(c.value).strip()]
        for i, (_, val) in enumerate(rv):
            for kw, field in kws.items():
                if kw in val.lower() and info[field] is None:
                    after = val.split(":",1)[1].strip() if ":" in val else ""
                    if after: info[field] = after
                    else:
                        for _, nv in rv[i+1:]:
                            if nv and nv not in (":",""):
                                info[field] = nv; break
    info["Năm học"] = normalize_nam_hoc(info["Năm học"])
    info["Học kỳ"] = normalize_hoc_ky(info["Học kỳ"])
    return info

def _process_formula_file(wb_raw, fname, file_bytes):
    """
    Đọc file dạng công thức. Chiến lược:
    1. Thử đọc sheet 'Điểm tổng hợp'/'diem_th' với data_only=True (cached value)
    2. Nếu không có cached value → đọc từ DSHVDT + diem_tx + diem_kt trực tiếp
    """
    sheets_raw = {s.lower(): wb_raw[s] for s in wb_raw.sheetnames}

    # ── Chiến lược 1: cached value trong sheet tổng hợp ───────
    ws_th_name = None
    for name in wb_raw.sheetnames:
        nl = name.lower()
        if "tổng hợp" in nl or "tong hop" in nl or "diem_th" in nl:
            ws_th_name = name; break

    if ws_th_name and file_bytes:
        try:
            wb_val = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws_val = wb_val[ws_th_name]
            info    = extract_general_info(ws_val)
            records = extract_score_table(ws_val)
            if records:
                return [{"File":fname,**info,**r} for r in records], None
        except: pass

    # ── Chiến lược 2: đọc từ DSHVDT + diem_tx + diem_kt ──────
    ws_ds = sheets_raw.get("dshvdt")
    ws_tx = sheets_raw.get("diem_tx")
    ws_kt = sheets_raw.get("diem_kt")
    if not ws_ds: return [], "Không tìm thấy sheet DSHVDT và không có cached value"

    info = _extract_info_dshvdt(ws_ds)
    _, data_start = _find_stt_data_start(ws_ds)
    if data_start is None: return [], "Không tìm thấy dữ liệu học viên trong DSHVDT"

    hr_ds, _ = _find_stt_data_start(ws_ds)
    mssv_col = ho_col = ten_col = hoten_col = None
    if hr_ds:
        for cell in ws_ds[hr_ds]:
            if cell.value and not is_formula(cell.value):
                t = str(cell.value).lower().strip()
                if "mssv" in t or "mshv" in t: mssv_col  = cell.column
                elif "họ và tên" in t:          hoten_col = cell.column
                elif "họ" in t:                 ho_col    = cell.column
                elif t in ("tên","ten"):         ten_col   = cell.column
    mssv_col = mssv_col or 3
    ho_col   = ho_col   or 4
    ten_col  = ten_col  or 5

    tx_col = tx_start = None
    if ws_tx:
        _, tx_start = _find_stt_data_start(ws_tx)
        if tx_start:
            for cell in ws_tx[tx_start]:
                if cell.value and not is_formula(cell.value) and cell.column > 1:
                    try: float(str(cell.value).replace(",",".")); tx_col = cell.column; break
                    except: pass

    kt_col = kt_start = None
    if ws_kt:
        hr_kt, kt_start = _find_stt_data_start(ws_kt)
        if kt_start and hr_kt:
            for cell in ws_kt[hr_kt]:
                if cell.value and not is_formula(cell.value):
                    if "điểm thi" in str(cell.value).lower(): kt_col = cell.column; break
        if kt_col is None and kt_start:
            for cell in ws_kt[kt_start]:
                if cell.value and not is_formula(cell.value) and cell.column > 1:
                    try: float(str(cell.value).replace(",",".")); kt_col = cell.column; break
                    except: pass

    records = []
    for offset in range(200):
        row_idx = data_start + offset
        stt = real_val(ws_ds, row_idx, 1)
        if stt is None: continue
        try: int(float(str(stt)))
        except: break

        mssv = str(real_val(ws_ds, row_idx, mssv_col) or "").strip()
        if not mssv or mssv.lower() in ("nan","none",""): continue

        if hoten_col:
            ho    = str(real_val(ws_ds, row_idx, hoten_col) or "").strip()
            extra = str(real_val(ws_ds, row_idx, hoten_col+1) or "").strip()
            hoten = f"{ho} {extra}".strip()
        else:
            ho    = str(real_val(ws_ds, row_idx, ho_col)  or "").strip()
            ten   = str(real_val(ws_ds, row_idx, ten_col) or "").strip()
            hoten = f"{ho} {ten}".strip()

        diem_tx = normalize_score(real_val(ws_tx, tx_start+offset, tx_col)) \
                  if ws_tx and tx_col and tx_start else None
        diem_kt = normalize_score(real_val(ws_kt, kt_start+offset, kt_col)) \
                  if ws_kt and kt_col and kt_start else None

        diem_tb = None
        if diem_tx is not None and diem_kt is not None:
            diem_tb = round(diem_tx * 0.3 + diem_kt * 0.7, 2)
        elif diem_kt is not None:
            diem_tb = diem_kt

        records.append({"MSSV":mssv,"Họ và tên":hoten,
            "Điểm quá trình (30%)":diem_tx,
            "Điểm giữa kỳ (30%)":  diem_tx,
            "Điểm thi (70%)":      diem_kt,
            "Điểm TB":             diem_tb,
            "Điểm làm tròn":       round_0_5(diem_tb),
            "Điểm chữ":            ""})

    return [{"File":fname,**info,**r} for r in records], None


def process_excel_file(file_bytes, fname):
    """
    Tự động nhận diện 2 loại file:
    - Loại 1 (value-based): có sheet 'Điểm tổng hợp' với giá trị thực
    - Loại 2 (formula-based): có sheet DSHVDT + diem_tx + diem_kt dùng công thức
    """
    # Đọc với data_only=False để nhận diện loại file và xử lý formula-based
    try: wb_raw = load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception as e: return [], str(e)

    sheets_lower = [s.lower() for s in wb_raw.sheetnames]

    # Loại 2: formula-based — có sheet DSHVDT hoặc sheet Điểm tổng hợp dùng công thức
    # Detect bằng cách kiểm tra sheet tổng hợp có công thức ở cột họ/tên không
    is_formula_based = "dshvdt" in sheets_lower
    if not is_formula_based:
        # Kiểm tra thêm: sheet tổng hợp có công thức không
        for name in wb_raw.sheetnames:
            nl = name.lower()
            if "tổng hợp" in nl or "tong hop" in nl or "diem_th" in nl:
                ws_chk = wb_raw[name]
                hr = find_header_row(ws_chk)
                if hr:
                    for r in range(hr+1, hr+4):
                        for c in ws_chk.iter_cols(min_row=r, max_row=r, min_col=1, max_col=10):
                            for cell in c:
                                if is_formula(cell.value):
                                    is_formula_based = True; break
                break

    if is_formula_based:
        return _process_formula_file(wb_raw, fname, file_bytes)

    # Loại 1: sheet Điểm tổng hợp với giá trị thực (đã save bằng Excel)
    try: wb_val = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e: return [], str(e)

    ws_val = None
    for name in wb_val.sheetnames:
        nl = name.lower()
        if "tổng hợp" in nl or "tong hop" in nl or "diem_th" in nl:
            ws_val = wb_val[name]; break
    if ws_val is None:
        return [], f"Không nhận dạng được loại file. Các sheet: {wb_raw.sheetnames}"

    info    = extract_general_info(ws_val)
    records = extract_score_table(ws_val)
    if not records:
        return [], f"Sheet '{ws_val.title}' không có dữ liệu học viên."
    return [{"File":fname,**info,**r} for r in records], None

@st.cache_data(show_spinner=False)
def load_ketqua_file(filepath):
    try: df=pd.read_excel(filepath,engine="openpyxl")
    except Exception as e: return None,str(e)
    missing=[c for c in ["Tên môn học","MSMH"] if c not in df.columns]
    if missing: return None,f"File thiếu cột: {missing}"
    df=(df.dropna(subset=["Tên môn học","MSMH"])
          .drop_duplicates(subset="Tên môn học")
          .reset_index(drop=True))
    return df,None

def check_data_files():
    present,missing=[],[]
    for cn,fname in CHUYEN_NGANH.items():
        (present if os.path.exists(os.path.join(DATA_DIR,fname)) else missing).append((cn,fname))
    return present,missing

def normalize_text(s):
    t = str(s).strip().lower()
    t = re.sub(r"[-\u2013\u2014]", " ", t)   # bỏ dấu gạch ngang/gạch nối
    t = t.replace("đ", "d")
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")  # bỏ hết dấu thanh điệu
    t = re.sub(r"\s+", " ", t).strip()
    return t

def merge_and_build(df_diem, df_mapping):
    mapping=df_mapping.copy()
    mapping["_key"]=mapping["Tên môn học"].apply(normalize_text)
    diem=df_diem.copy()
    diem["_key"]=diem["Tên môn học"].apply(normalize_text)
    df_merged=diem.merge(mapping.drop(columns=["Tên môn học"]),on="_key",how="left").drop(columns=["_key"])
    def mlhp(row):
        nam2=str(row["Năm học"]).split("-")[0][-2:]
        hk=HK_MAP.get(str(row["Học kỳ"]).strip(),str(row["Học kỳ"]).strip())
        msmh=str(row["MSMH"]) if pd.notna(row["MSMH"]) else "???"
        return f"{nam2}{hk}{msmh}01"
    df_merged["Mã lớp học phần"]=df_merged.apply(mlhp,axis=1)
    df_out=df_merged.rename(columns={"Lớp":"Mã lớp"})
    # Sheet Daydu
    want=["Năm học","Học kỳ","Khoa","Mã lớp","Tên môn học","Số tín chỉ","MSMH","Mã lớp học phần",
          "MSSV","Họ và tên","Điểm quá trình (30%)","Điểm giữa kỳ (30%)", "Điểm thi (70%)","Điểm TB","Điểm làm tròn","Điểm chữ",
          "Trọng số điểm QT","Trọng số điểm GK","Trọng số điểm CK","Điểm đạt"]
    df_daydu=df_out[[c for c in want if c in df_out.columns]]
    # Sheet LopHocPhan
    lhp_cols=[c for c in ["Năm học","Học kỳ","Mã lớp","MSMH","Mã lớp học phần","Tên môn học"] if c in df_out.columns]
    df_lhp=(df_out[lhp_cols].drop_duplicates(subset="Mã lớp học phần").reset_index(drop=True))
    df_lhp["Rớt môn nếu rớt điểm thành phần"]=""
    # Sheet SV
    sv_cols=[c for c in ["MSSV","Họ và tên","Mã lớp học phần"] if c in df_out.columns]
    df_sv=df_out[sv_cols].reset_index(drop=True)
    no_match=df_merged[df_merged["MSMH"].isna()]["Tên môn học"].dropna().unique().tolist()
    return df_daydu,df_lhp,df_sv,no_match

def to_excel_bytes(df_daydu,df_lhp,df_sv):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        df_daydu.to_excel(w,sheet_name="Daydu",index=False)
        df_lhp.to_excel(  w,sheet_name="LopHocPhan",index=False)
        df_sv.to_excel(   w,sheet_name="SV_LopHocPhan",index=False)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero">
  <h1>🎓 Xử lý Bảng Điểm Tổng Hợp</h1>
  <p>Upload file Excel bảng điểm &nbsp;·&nbsp; Chọn chuyên ngành &nbsp;·&nbsp; Tải kết quả (3 sheet)</p>
</div>
""", unsafe_allow_html=True)

# Kiểm tra file data
present_files, missing_files = check_data_files()
available_cn = [cn for cn,_ in present_files]

if missing_files:
    with st.expander(f"⚠ {len(missing_files)} chuyên ngành chưa có file — xem chi tiết"):
        st.markdown(
            '<div class="miss">Đặt các file sau vào thư mục <code>data/</code>:<br><br>' +
            "<br>".join(f"&nbsp;• <code>{f}</code> &nbsp;({cn})" for cn,f in missing_files) +
            "</div>", unsafe_allow_html=True)

if not available_cn:
    st.markdown('<div class="err">❌ Chưa có file dữ liệu nào trong thư mục <code>data/</code>.</div>',
                unsafe_allow_html=True)
    st.stop()

# BƯỚC 1
st.markdown('<div class="card"><div class="step">Bước 1 — Upload file bảng điểm</div>',
            unsafe_allow_html=True)
st.caption("Chọn 1 hoặc nhiều file Excel — mỗi file cần có sheet **Điểm tổng hợp**.")
uploaded_files = st.file_uploader(
    label="upload", type=["xlsx","xls"],
    accept_multiple_files=True, label_visibility="collapsed")
if uploaded_files:
    st.caption("✔ Đã chọn **{}** file: {}".format(
        len(uploaded_files), ", ".join(f"`{f.name}`" for f in uploaded_files)))
st.markdown("</div>", unsafe_allow_html=True)

# BƯỚC 2
st.markdown('<div class="card"><div class="step">Bước 2 — Chọn chuyên ngành</div>',
            unsafe_allow_html=True)
selected_cn = st.selectbox("cn", options=available_cn, label_visibility="collapsed")
if selected_cn:
    fpath = os.path.join(DATA_DIR, CHUYEN_NGANH[selected_cn])
    df_check, err_check = load_ketqua_file(fpath)
    if df_check is not None:
        st.caption(f"📂 `{CHUYEN_NGANH[selected_cn]}` · **{len(df_check)}** môn học")
    else:
        st.markdown(f'<div class="err">❌ {err_check}</div>', unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)

# BƯỚC 3
st.markdown('<div class="card"><div class="step">Bước 3 — Xử lý & Tải kết quả</div>',
            unsafe_allow_html=True)

if st.button("▶  Xử lý ngay", use_container_width=True, type="primary"):
    if not uploaded_files:
        st.markdown('<div class="warn">⚠ Chưa upload file bảng điểm nào.</div>',
                    unsafe_allow_html=True)
    else:
        with st.spinner("Đang xử lý…"):
            fpath = os.path.join(DATA_DIR, CHUYEN_NGANH[selected_cn])
            df_mapping, map_err = load_ketqua_file(fpath)

            if map_err:
                st.markdown(f'<div class="err">❌ {map_err}</div>', unsafe_allow_html=True)
            else:
                all_rows=[]; file_errors=[]; file_stats=[]
                for uf in uploaded_files:
                    rows,err=process_excel_file(uf.read(), uf.name)
                    if err: file_errors.append((uf.name,err))
                    else:   all_rows.extend(rows); file_stats.append((uf.name,len(rows)))

                for fn,e in file_errors:
                    st.markdown(f'<div class="warn">⚠ <b>{fn}</b>: {e}</div>',
                                unsafe_allow_html=True)

                if not all_rows:
                    st.markdown('<div class="err">❌ Không trích xuất được dữ liệu.</div>',
                                unsafe_allow_html=True)
                else:
                    df_diem=pd.DataFrame(all_rows)
                    df_daydu,df_lhp,df_sv,no_match=merge_and_build(df_diem,df_mapping)
                    excel_bytes=to_excel_bytes(df_daydu,df_lhp,df_sv)

                    st.markdown(f"""
                    <div class="ok">
                      <h4>✅ Hoàn tất — {selected_cn}</h4>
                      📄 <b>{len(file_stats)}</b> file &nbsp;·&nbsp;
                      👥 <b>{len(df_daydu)}</b> học viên &nbsp;·&nbsp;
                      📚 <b>{len(df_lhp)}</b> lớp học phần
                    </div>""", unsafe_allow_html=True)

                    if len(file_stats)>1:
                        with st.expander("Chi tiết từng file"):
                            for fn,cnt in file_stats:
                                st.write(f"• `{fn}` — {cnt} học viên")

                    if no_match:
                        st.markdown(
                            '<div class="warn">⚠ <b>Không tìm được MSMH cho '
                            f'{len(no_match)} môn</b> (tên môn trong bảng điểm '
                            'không khớp với file chuyên ngành):<br>'
                            +"<br>".join(f"&nbsp;&nbsp;• {m}" for m in no_match)
                            +"</div>", unsafe_allow_html=True)

                    with st.expander("👁 Xem trước kết quả", expanded=True):
                        t1,t2,t3=st.tabs(["📋 Daydu","🏫 LopHocPhan","👥 SV_LopHocPhan"])
                        with t1: st.dataframe(df_daydu, use_container_width=True, height=300)
                        with t2: st.dataframe(df_lhp,   use_container_width=True)
                        with t3: st.dataframe(df_sv,    use_container_width=True, height=300)

                    st.download_button(
                        label="⬇  Tải file Excel kết quả (3 sheet)",
                        data=excel_bytes,
                        file_name=f"KetQua_{selected_cn.replace(' ','_').replace('-','')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)

st.markdown("</div>", unsafe_allow_html=True)
st.markdown("<br><center style='color:#94a3b8;font-size:.76rem'>"
            "Bảng Điểm Tổng Hợp · Đại học Khoa học Sức khỏe</center>",
            unsafe_allow_html=True)
